import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import PieChart, BarChart, Reference
import tkinter as tk
from tkinter import filedialog, messagebox
import os
from datetime import datetime
import re
from pandas import isna

# Obter o ano atual
CURRENT_YEAR = datetime.now().year

# Definir estilo de borda fina
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Definir cores
header_fill = PatternFill(start_color='EC7233', end_color='EC7233', fill_type='solid')  # Laranja
total_fill = PatternFill(start_color='D0CECE', end_color='D0CECE', fill_type='solid')   # Cinza

# Função para limpar e padronizar valores monetários
def clean_monetary_value(value):
    if pd.isna(value) or value == '':
        return 0.0
    try:
        cleaned = str(value).replace('R$', '').replace(' ', '').replace('.', '').replace(',', '.')
        return float(cleaned)
    except (ValueError, TypeError):
        return 0.0

# Função para limpar e padronizar valores de data
def clean_date_value(date_value):
    if isna(date_value) or date_value is None or date_value == '':
        return None
    if isinstance(date_value, str) and date_value.strip():
        date_str = date_value.strip()
        try:
            parsed_date = datetime.strptime(date_str, '%d/%m/%Y')
        except ValueError:
            try:
                parsed_date = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
            except ValueError:
                print(f"Erro ao parsear data '{date_str}'")
                return None
        if parsed_date.year == 1901:
            parsed_date = parsed_date.replace(year=CURRENT_YEAR)
        return parsed_date
    if isinstance(date_value, (int, float)):
        try:
            parsed_date = pd.to_datetime('1899-12-30') + pd.to_timedelta(date_value, unit='D')
            if parsed_date.year == 1901:
                parsed_date = parsed_date.replace(year=CURRENT_YEAR)
            return parsed_date
        except (ValueError, TypeError):
            return None
    print(f"Valor de data inválido: {date_value} (tipo: {type(date_value)})")
    return None

# Função para carregar e limpar dados de CMCL904-CLIENTE-CC.xlsx
def load_client_data(file_path):
    try:
        # Carrega a primeira aba do arquivo
        df = pd.read_excel(file_path, sheet_name=0, dtype={'Emissão': str, 'IDA': str, 'VOLTA': str})
        df.columns = df.columns.str.strip()
        
        # Lista de colunas esperadas
        required_columns = ['Razão Social', 'cnpj', 'Centro de Custo', 'Fornecedor', 'Tarifas', 
                           'Tx.Embq.', 'Tx.Serviço', 'Total', 'Passageiro', 'Solicitante', 
                           'Documento', 'Trecho', 'Emissão', 'IDA', 'VOLTA']
        
        # Verifica se todas as colunas necessárias estão presentes
        missing_columns = [col for col in required_columns if col not in df.columns]
        if missing_columns:
            raise ValueError(f"Colunas ausentes no arquivo: {', '.join(missing_columns)}")
        
        # Limpa valores monetários
        monetary_cols = ['Tarifas', 'Tx.Embq.', 'Tx.Serviço', 'Tx.Extra', 'Total', 'Valor Medio']
        for col in monetary_cols:
            if col in df.columns:
                df[col] = df[col].apply(clean_monetary_value)
        
        # Limpa valores de data
        df['Emissão'] = df['Emissão'].apply(clean_date_value)
        df['IDA'] = df['IDA'].apply(clean_date_value)
        df['VOLTA'] = df['VOLTA'].apply(clean_date_value)
        
        # Filtra linhas de totais/subtotais
        df = df[~df['Razão Social'].str.contains('Total|Subtotal', na=False, case=False, regex=True)]
        df = df[~df['Trecho'].str.contains('Total|Subtotal', na=False, case=False, regex=True)]
        
        return df
    except FileNotFoundError:
        messagebox.showerror("Erro", f"Arquivo não encontrado: {file_path}")
        return None
    except ValueError as ve:
        messagebox.showerror("Erro", f"Falha ao carregar dados do cliente: {str(ve)}")
        return None
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao carregar dados do cliente: {str(e)}. Verifique o formato do arquivo e o nome das colunas.")
        return None

# Função para carregar e limpar dados de CMCL904-FORNECEDOR.xlsx
def load_supplier_data(file_path):
    try:
        df = pd.read_excel(file_path, sheet_name=0)
        df.columns = df.columns.str.strip()
        monetary_cols = ['Tarifas', 'Tx.Embq.', 'Tx.Serviço', 'Tx.Extra', 'Total', 'Valor Medio']
        for col in monetary_cols:
            if col in df.columns:
                df[col] = df[col].apply(clean_monetary_value)
        df = df[~df['Fornecedor'].str.contains('Total', na=False, case=False)]
        return df
    except FileNotFoundError:
        messagebox.showerror("Erro", f"Arquivo não encontrado: {file_path}")
        return None
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao carregar dados do fornecedor: {str(e)}")
        return None

# Função para criar a aba EMISSOES
def create_emissoes_sheet(client_df, workbook):
    ws = workbook.create_sheet("EMISSOES")
    headers = ['RAZAO SOC', 'CNPJ', 'CENTRO DE CUSTO', 'CIA', 'TARIFA', 'TAXA DE EMBARQUE',
               'TAXA DE SERVIÇO', 'TOTAL', 'VIAJANTE', 'SOLICITANTE', 'LOCALIZADOR BILHETE',
               'TRECHO COMPL.', 'DT. EMISSAO', 'DT. PARTIDA', 'DT. RETORNO']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border
    
    data_rows = []
    for _, row in client_df.iterrows():
        cnpj = row['cnpj'] if pd.notna(row['cnpj']) and row['cnpj'] != '' else "Não Informado"
        localizador = row['Documento'] if pd.notna(row['Documento']) and row['Documento'] != '' else "Não Informado"
        dt_emissao = row['Emissão']
        dt_partida = row['IDA']
        dt_retorno = row['VOLTA']
        
        data_rows.append([
            row['Razão Social'],
            cnpj,
            row['Centro de Custo'] if pd.notna(row['Centro de Custo']) else 'A DEFINIR',
            row['Fornecedor'],
            row['Tarifas'],
            row['Tx.Embq.'],
            row['Tx.Serviço'],
            row['Total'],
            row['Passageiro'],
            row['Solicitante'],
            localizador,
            row['Trecho'],
            dt_emissao,
            dt_partida,
            dt_retorno
        ])
    
    for row_idx, data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            if isinstance(value, datetime):
                cell.number_format = 'DD/MM/YYYY'
                cell.alignment = Alignment(horizontal='center')
            elif isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
    
    if data_rows:
        total_row = ['Total Geral'] + [''] * (len(headers) - 1)
        totals = {'TARIFA': 0, 'TAXA DE EMBARQUE': 0, 'TAXA DE SERVIÇO': 0, 'TOTAL': 0}
        for row in data_rows:
            totals['TARIFA'] += row[4] if isinstance(row[4], (int, float)) else 0
            totals['TAXA DE EMBARQUE'] += row[5] if isinstance(row[5], (int, float)) else 0
            totals['TAXA DE SERVIÇO'] += row[6] if isinstance(row[6], (int, float)) else 0
            totals['TOTAL'] += row[7] if isinstance(row[7], (int, float)) else 0
        total_row[4] = totals['TARIFA']
        total_row[5] = totals['TAXA DE EMBARQUE']
        total_row[6] = totals['TAXA DE SERVIÇO']
        total_row[7] = totals['TOTAL']
        for col_idx, value in enumerate(total_row, 1):
            cell = ws.cell(row=len(data_rows) + 2, column=col_idx)
            cell.value = value
            cell.border = thin_border
            cell.font = Font(bold=True)
            cell.fill = total_fill
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    
    return ws

# Função para criar a aba EMISSÃO E REEMISSAO
def create_emissao_reemissao_sheet(client_df, workbook):
    ws = workbook.create_sheet("EMISSÃO E REEMISSAO")
    headers = ['EMISSÃO/REMISSÃO', 'VALOR TARIFA', 'VALOR TAXAS',
               'Total QUANTIDADE DE BILHETES', 'Valor Total', 'TICKET MÉDIO', 'PERCENTUAL %']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border
    
    total_bilhetes = len(client_df)
    total_tarifa = client_df['Tarifas'].sum()
    total_taxas = client_df['Tx.Embq.'].sum() + client_df['Tx.Serviço'].sum()
    total_valor = client_df['Total'].sum()
    #ticket_medio = total_valor / total_bilhetes if total_bilhetes > 0 else 0
    #Israel Ruiz 03/07/2025 - Calculo pela tarifa 
    ticket_medio = total_tarifa / total_bilhetes if total_bilhetes > 0 else 0
    data_emissao = [
        ['EMISSÃO',  total_tarifa, total_taxas, total_bilhetes, total_valor, ticket_medio, 100.0],
        ['REMISSAO', 0, 0, 0, 0, 0, 0],
        ['TOTAL',  total_tarifa, total_taxas, total_bilhetes, total_valor, ticket_medio, 100.0]
    ]
    
    for row_idx, data in enumerate(data_emissao, 2):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            if row_idx == len(data_emissao) + 1:
                cell.font = Font(bold=True)
                cell.fill = total_fill
            if isinstance(value, (int, float)) and col_idx not in [1]:
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

# Função para criar a aba TOTAL POR EMPRESAS
def create_empresa_sheet(client_df, workbook):
    ws = workbook.create_sheet("TOTAL POR EMPRESAS")
    headers = ['EMPRESA', 'TOTAL', 'PERCENTUAL %']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border
    
    empresa_totals = client_df.groupby('Razão Social')['Total'].sum().reset_index()
    total_geral = empresa_totals['Total'].sum()
    data_empresa = []
    for _, row in empresa_totals.iterrows():
        percentual = (row['Total'] / total_geral)*100 if total_geral > 0 else 0
        data_empresa.append([row['Razão Social'], row['Total'], percentual])
    data_empresa.append(['TOTAL', total_geral, 100.0])
    
    for row_idx, data in enumerate(data_empresa, 2):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            if row_idx == len(data_empresa) + 1:
                cell.font = Font(bold=True)
                cell.fill = total_fill
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

# Função para criar a aba TOTAL POR CENTRO DE CUSTO
def create_centro_custo_sheet(client_df, workbook):
    ws = workbook.create_sheet("TOTAL POR CENTRO DE CUSTO")
    headers = ['CENTRO DE CUSTO', 'QUANTIDADE DE BILHETE', 'VALOR TOTAL', 'PERCENTUAL %']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border
    
    centro_totals = client_df.groupby('Centro de Custo').agg({'Total': 'sum', 'Razão Social': 'count'}).reset_index()
    centro_totals.columns = ['Centro de Custo', 'Total', 'Quantidade']
    total_bilhetes = centro_totals['Quantidade'].sum()
    total_valor = centro_totals['Total'].sum()
    data_centro = []
    for _, row in centro_totals.iterrows():
        percentual = (row['Total'] / total_valor)*100 if total_valor > 0 else 0
        data_centro.append([row['Centro de Custo'], row['Quantidade'], row['Total'], percentual])
    data_centro.append(['TOTAL', total_bilhetes, total_valor, 100.0])
    
    for row_idx, data in enumerate(data_centro, 2):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            if row_idx == len(data_centro) + 1:
                cell.font = Font(bold=True)
                cell.fill = total_fill
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

# Função para criar a aba TOTAL POR CIA AEREA
def create_cia_aerea_sheet(client_df, workbook):
    ws = workbook.create_sheet("TOTAL POR CIA AEREA")
    headers = ['CIA NAC', 'QUANTIDADE DE BILHETES', 'VALOR DA TARIFA', 'VALOR TAXAS',
               'Valor Total', 'TICKET MÉDIO', 'PERCENTUAL %']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border
    
    cia_totals = client_df.groupby('Fornecedor').agg({
        'Total': 'sum',
        'Tarifas': 'sum',
        'Tx.Embq.': 'sum',
        'Tx.Serviço': 'sum',
        'Razão Social': 'count'
    }).reset_index()
    cia_totals['Taxas'] = cia_totals['Tx.Embq.'] + cia_totals['Tx.Serviço']
    cia_totals.columns = ['CIA NAC', 'Total', 'VALOR DA TARIFA', 'Tx.Embq.', 'Tx.Serviço', 'QUANTIDADE DE BILHETES', 'VALOR TAXAS']
    total_valor = cia_totals['Total'].sum()
    total_bilhetes = cia_totals['QUANTIDADE DE BILHETES'].sum()
    total_tarifa = cia_totals['VALOR DA TARIFA'].sum()
    total_taxas = cia_totals['VALOR TAXAS'].sum()
    data_cia = []
    for _, row in cia_totals.iterrows():
        #ticket_medio = row['Total'] / row['QUANTIDADE DE BILHETES'] if row['QUANTIDADE DE BILHETES'] > 0 else 0
        ticket_medio = row['VALOR DA TARIFA'] / row['QUANTIDADE DE BILHETES'] if row['QUANTIDADE DE BILHETES'] > 0 else 0
        percentual = (row['Total'] / total_valor)*100 if total_valor > 0 else 0
        data_cia.append([
            row['CIA NAC'], row['QUANTIDADE DE BILHETES'], row['VALOR DA TARIFA'],
            row['VALOR TAXAS'], row['Total'], ticket_medio, percentual
        ])
    data_cia.append(['TOTAL', total_bilhetes, total_tarifa, total_taxas, total_valor, ticket_medio, 100.0])
    
    for row_idx, data in enumerate(data_cia, 2):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            if row_idx == len(data_cia) + 1:
                cell.font = Font(bold=True)
                cell.fill = total_fill
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width
    
    if len(data_cia) > 1:
        client_df['Mês'] = client_df['Emissão'].apply(lambda x: x.strftime('%m/%Y') if pd.notna(x) else 'Sem Mês')
        pivot_data = client_df.pivot_table(
            values='Tarifas',
            index='Fornecedor',
            columns='Mês',
            aggfunc='sum',
            fill_value=0
        ).reset_index()
        
        pivot_headers = ['Fornecedor'] + list(pivot_data.columns[1:])
        pivot_start_row = len(data_cia) + 4
        for col_idx, header in enumerate(pivot_headers, 1):
            cell = ws.cell(row=pivot_start_row, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_idx, row in enumerate(pivot_data.itertuples(index=False), pivot_start_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = thin_border
                if col_idx > 1:
                    cell.number_format = '#,##0.00'
                    cell.alignment = Alignment(horizontal='right')
                else:
                    cell.alignment = Alignment(horizontal='left')
        
        # Posiciona o gráfico de barras duas linhas abaixo da tabela dinâmica, na coluna B
        chart_row = pivot_start_row + len(pivot_data) + 2  # Tabela dinâmica termina em pivot_start_row + len(pivot_data), +2 linhas
        bar_chart = BarChart()
        bar_chart.title = "Tarifas por Companhia Aérea e Mês"
        bar_data = Reference(ws, min_col=2, min_row=pivot_start_row + 1, max_col=len(pivot_headers), max_row=pivot_start_row + len(pivot_data))
        bar_cats = Reference(ws, min_col=1, min_row=pivot_start_row + 1, max_row=pivot_start_row + len(pivot_data))
        bar_chart.add_data(bar_data, titles_from_data=True)
        bar_chart.set_categories(bar_cats)
        bar_chart.x_axis.title = "Companhia Aérea"
        bar_chart.y_axis.title = "Valor das Tarifas"
        bar_chart.height = 8
        bar_chart.width = 12
        ws.add_chart(bar_chart, f"B{chart_row}")
    
    if len(data_cia) > 1:
        # Posiciona o gráfico de pizza ao lado do gráfico de barras, na coluna I
        pie_chart = PieChart()
        pie_chart.title = "Percentual por Companhia Aérea"
        pie_data = Reference(ws, min_col=7, min_row=2, max_row=len(data_cia))
        pie_cats = Reference(ws, min_col=1, min_row=2, max_row=len(data_cia))
        pie_chart.add_data(pie_data, titles_from_data=False)
        pie_chart.set_categories(pie_cats)
        pie_chart.height = 8
        pie_chart.width = 12
        ws.add_chart(pie_chart, f"F{chart_row}")
    
    return ws

# Função para criar a aba TOTAL POR CIA E TRECHO
def create_cia_trecho_sheet(client_df, workbook):
    ws = workbook.create_sheet("TOTAL POR CIA E TRECHO")
    headers = ['CIA', 'TRECHO', 'QUANTIDADE DE BILHETES', 'VALOR DA TARIFA',
               'VALOR DAS TAXAS', 'VALOR TOTAL', 'TICKET MÉDIO', 'PERCENTUAL %']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border
    
    cia_trecho_totals = client_df.groupby(['Fornecedor', 'Trecho']).agg({
        'Total': 'sum',
        'Tarifas': 'sum',
        'Tx.Embq.': 'sum',
        'Tx.Serviço': 'sum',
        'Razão Social': 'count'
    }).reset_index()
    cia_trecho_totals['Taxas'] = cia_trecho_totals['Tx.Embq.'] + cia_trecho_totals['Tx.Serviço']
    cia_trecho_totals.columns = ['CIA', 'TRECHO', 'Total', 'VALOR DA TARIFA', 'Tx.Embq.', 'Tx.Serviço', 'QUANTIDADE DE BILHETES', 'VALOR DAS TAXAS']
    total_valor = cia_trecho_totals['Total'].sum()
    total_bilhetes = cia_trecho_totals['QUANTIDADE DE BILHETES'].sum()
    total_tarifa = cia_trecho_totals['VALOR DA TARIFA'].sum()
    total_taxas = cia_trecho_totals['VALOR DAS TAXAS'].sum()
    data_cia_trecho = []
    for _, row in cia_trecho_totals.iterrows():
        #ticket_medio = row['Total'] / row['QUANTIDADE DE BILHETES'] if row['QUANTIDADE DE BILHETES'] > 0 else 0
        ticket_medio = row['VALOR DA TARIFA'] / row['QUANTIDADE DE BILHETES'] if row['QUANTIDADE DE BILHETES'] > 0 else 0
        percentual = (row['Total'] / total_valor)*100 if total_valor > 0 else 0
        data_cia_trecho.append([
            row['CIA'], row['TRECHO'], row['QUANTIDADE DE BILHETES'], row['VALOR DA TARIFA'],
            row['VALOR DAS TAXAS'], row['Total'], ticket_medio, percentual
        ])
    data_cia_trecho.append(['TOTAL', '', total_bilhetes, total_tarifa, total_taxas, total_valor, ticket_medio, 100.0])
    
    for row_idx, data in enumerate(data_cia_trecho, 2):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            if row_idx == len(data_cia_trecho) + 1:
                cell.font = Font(bold=True)
                cell.fill = total_fill
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

# Função para criar a aba TOTAL POR SOLICITANTE
def create_solicitante_sheet(client_df, workbook):
    ws = workbook.create_sheet("TOTAL POR SOLICITANTE")
    headers = ['SOLICITANTE', 'VALOR TOTAL', 'PERCENTUAL %']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border
    
    solicitante_totals = client_df.groupby('Solicitante')['Total'].sum().reset_index()
    total_valor = solicitante_totals['Total'].sum()
    data_solicitante = []
    for _, row in solicitante_totals.iterrows():
        percentual = (row['Total'] / total_valor)*100 if total_valor > 0 else 0
        data_solicitante.append([row['Solicitante'], row['Total'], percentual])
    data_solicitante.append(['TOTAL', total_valor, 100.0])
    
    for row_idx, data in enumerate(data_solicitante, 2):
        for col_idx, value in enumerate(data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            if row_idx == len(data_solicitante) + 1:
                cell.font = Font(bold=True)
                cell.fill = total_fill
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.alignment = Alignment(horizontal='left')
    
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

# Função para criar a aba TOTAL CREDITOS DISPONIVEIS
def create_creditos_disponiveis_sheet(client_df, workbook):
    ws = workbook.create_sheet("TOTAL CREDITOS DISPONIVEIS")
    headers = ['PASSAGEIRO', 'CIA', 'LOCALIZADOR', 'VALOR DA TARIFA', 'VALOR TAXAS', 'VALOR TOTAL', 'DISPONIVEL']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.fill = header_fill
        cell.border = thin_border
    
    # Ajusta a largura das colunas com base nos cabeçalhos
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

# Função principal para processar arquivos e gerar saída
def process_files(client_file, supplier_file, output_file):
    client_df = load_client_data(client_file)
    supplier_df = load_supplier_data(supplier_file)
    
    if client_df is None or supplier_df is None:
        return
    
    wb = openpyxl.Workbook()
    wb.remove(wb['Sheet'])
    
    create_emissoes_sheet(client_df, wb)
    create_emissao_reemissao_sheet(client_df, wb)
    create_empresa_sheet(client_df, wb)
    create_centro_custo_sheet(client_df, wb)
    create_cia_aerea_sheet(client_df, wb)
    create_cia_trecho_sheet(client_df, wb)
    create_solicitante_sheet(client_df, wb)
    create_creditos_disponiveis_sheet(client_df, wb)
    
    try:
        wb.save(output_file)
        messagebox.showinfo("Sucesso", f"Arquivo Excel gerado com sucesso em {output_file}")
    except Exception as e:
        messagebox.showerror("Erro", f"Falha ao salvar arquivo Excel: {str(e)}")

# Interface gráfica
class ExcelProcessorApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Gerador de Relatórios Excel")
        self.root.geometry("600x400")
        
        self.client_file = tk.StringVar()
        self.supplier_file = tk.StringVar()
        self.output_file = tk.StringVar()
        
        tk.Label(root, text="Arquivo com dados cliente:").grid(row=0, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(root, textvariable=self.client_file, width=50).grid(row=0, column=1, padx=5, pady=5)
        tk.Button(root, text="Selecionar", command=self.browse_client_file).grid(row=0, column=2, padx=5, pady=5)
        
        tk.Label(root, text="Arquivo com dados FORNECEDOR:").grid(row=1, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(root, textvariable=self.supplier_file, width=50).grid(row=1, column=1, padx=5, pady=5)
        tk.Button(root, text="Selecionar", command=self.browse_supplier_file).grid(row=1, column=2, padx=5, pady=5)
        
        tk.Label(root, text="Arquivo de Saída:").grid(row=2, column=0, padx=5, pady=5, sticky='w')
        tk.Entry(root, textvariable=self.output_file, width=50).grid(row=2, column=1, padx=5, pady=5)
        tk.Button(root, text="Selecionar", command=self.browse_output_file).grid(row=2, column=2, padx=5, pady=5)
        
        tk.Button(root, text="Gerar Relatório", command=self.generate_report).grid(row=3, column=0, padx=5, pady=10)
        tk.Button(root, text="Sair", command=self.exit_app).grid(row=3, column=2, padx=5, pady=10)
    
    def browse_client_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx *.xls")])
        if file_path:
            self.client_file.set(file_path)
    
    def browse_supplier_file(self):
        file_path = filedialog.askopenfilename(filetypes=[("Arquivos Excel", "*.xlsx *.xls")])
        if file_path:
            self.supplier_file.set(file_path)
    
    def browse_output_file(self):
        file_path = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Arquivos Excel", "*.xlsx")])
        if file_path:
            self.output_file.set(file_path)
    
    def generate_report(self):
        if not self.client_file.get() or not self.supplier_file.get() or not self.output_file.get():
            messagebox.showwarning("Aviso", "Por favor, selecione todos os arquivos de entrada e saída.")
            return
        process_files(self.client_file.get(), self.supplier_file.get(), self.output_file.get())
    
    def exit_app(self):
        self.root.destroy()

# Executar a aplicação
if __name__ == "__main__":
    root = tk.Tk()
    app = ExcelProcessorApp(root)
    root.mainloop()